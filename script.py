import requests
from datetime import datetime, date

from time import time, sleep
import multiprocessing
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter
import os

headers = {
        'Connection': 'keep-alive',
        'Cache-Control': 'max-age=0',
        'DNT': '1',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.79 Safari/537.36',
        'Sec-Fetch-User': '?1',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Sec-Fetch-Site': 'none',
        'Sec-Fetch-Mode': 'navigate',
        'Accept-Language': 'en-US,en;q=0.9,hi;q=0.8',
    }

def nsefetch(payload):

    try:
        output = requests.get(payload,headers=headers).json()
        return output[-1]['nope']

    except ValueError:
        s =requests.Session()
        output = s.get("https://nopechart.com",headers=headers)
        output = s.get(payload,headers=headers)
        return output[-1]['nope']


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,truncate_sheet=False, **to_excel_kwargs):
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

def func(symbol,ts):
    today = date.today()
    d1 = today.strftime("%m-%d-%Y")
    nope = nsefetch(f'https://nopechart.com/cache/{symbol}_{d1}.json?_={ts}')
    
    date_time = datetime.fromtimestamp(ts)
    ct = date_time.strftime("%H:%M:%S")
    print(symbol + ' @ '+ct+' = '+str(nope))
    return nope
    
def init_file(file_name):
    df = pd.DataFrame(columns = ['Time', 'NOPE Value'])
    append_df_to_excel(f"{file_name}.xlsx",df, index=False, truncate_sheet = True)

stock = input('Enter the stock name : ').upper()

today = date.today()
d = today.strftime("%d-%m-%Y")
file_name = stock+' '+d
init_file(file_name)

while(True):
    try:
        
        current_time_stamp = int(datetime.now().timestamp())
        date_time = datetime.fromtimestamp(current_time_stamp)
        ct = date_time.strftime("%H:%M:%S")
        df = pd.DataFrame(columns = ['Time', 'NOPE Value'])
        df = df.append({'Time' : ct, 'NOPE Value' : func(stock,current_time_stamp)},ignore_index = True)
        append_df_to_excel(f"{file_name}.xlsx",df, header=None, index=False)
        sleep(20)
        
    except:
        print('Program Closed')
        break